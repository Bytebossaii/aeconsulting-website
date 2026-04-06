#!/usr/bin/env python3
"""
A&E Consulting LLC — Job Board Monitor
Scrapes StepStone, Indeed DE, and LinkedIn Jobs daily for target positions.
Emails a digest to enes@aeconsultingllc.de every morning.

SETUP:
  1. Fill in your Gmail credentials in CONFIG below (use App Password, not real password)
  2. Run manually:        python job_monitor_ae.py
  3. Run daily at 7AM:    Set up Windows Task Scheduler (instructions at bottom of file)

REQUIREMENTS:
  pip install requests beautifulsoup4
"""

import requests
from bs4 import BeautifulSoup
import smtplib
import json
import os
import time
import random
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
CONFIG = {
    # SMTP - Namecheap Private Email
    "smtp_host":       "mail.privateemail.com",
    "smtp_port":       587,
    "sender_email":    "enes@aeconsultingllc.de",
    "sender_password": "Reward.jabra25",
    "sender_name":     "Enes Haziri",
    "sender_title":    "Gruender",
    "company_name":    "A&E Consulting LLC",
    "website":         "www.aeconsultingllc.de",
    "phone":           "+383 49 677 019",
    "linkedin":        "",
    # Monitor settings
    "recipient_email": "enes@aeconsultingllc.de",
    "seen_jobs_file":  "seen_jobs.json",
    "min_delay":       2,
    "max_delay":       5,
    "excel_file":      "ae_job_leads.xlsx",
}

# ─────────────────────────────────────────────────────────────────────────────
# TARGET KEYWORDS — positions you want to find
# ─────────────────────────────────────────────────────────────────────────────
SEARCH_QUERIES = [
    # German property management
    "Sachbearbeiter Hausverwaltung",
    "Sachbearbeiter Immobilienverwaltung",
    "Mietsachbearbeiter",
    "Nebenkostenabrechnung",
    "Mietbuchhaltung remote",
    "Property Manager remote",
    "Immobilien Back Office",
    "Verwalter Assistent",
    # General back office German
    "Back Office Mitarbeiter remote",
    "Verwaltungsmitarbeiter remote",
    "Sachbearbeiter remote Deutsch",
    # English property management
    "property management back office remote",
    "lease administrator remote",
    "real estate admin remote",
    "property administrator remote",
]

# ─────────────────────────────────────────────────────────────────────────────
# HEADERS — mimic a real browser
# ─────────────────────────────────────────────────────────────────────────────
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ─────────────────────────────────────────────────────────────────────────────
# SEEN JOBS — prevents sending duplicates
# ─────────────────────────────────────────────────────────────────────────────
def load_seen_jobs():
    if os.path.exists(CONFIG["seen_jobs_file"]):
        with open(CONFIG["seen_jobs_file"], "r") as f:
            return set(json.load(f))
    return set()

def save_seen_jobs(seen):
    with open(CONFIG["seen_jobs_file"], "w") as f:
        json.dump(list(seen), f)

# ─────────────────────────────────────────────────────────────────────────────
# SCRAPER: StepStone.de
# ─────────────────────────────────────────────────────────────────────────────
def scrape_stepstone(query):
    jobs = []
    try:
        q = query.replace(" ", "+")
        url = f"https://www.stepstone.de/jobs/{q}?radius=0&remoteOption=remote"
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            return jobs
        soup = BeautifulSoup(resp.text, "html.parser")

        # StepStone job cards
        cards = soup.find_all("article", {"data-testid": "job-item"})
        if not cards:
            # Fallback selector
            cards = soup.find_all("article", class_=lambda x: x and "job" in x.lower())

        for card in cards[:8]:
            try:
                # Title
                title_el = card.find(["h2", "h3", "a"], {"data-testid": "job-item-title"})
                if not title_el:
                    title_el = card.find(["h2", "h3"])
                title = title_el.get_text(strip=True) if title_el else "N/A"

                # Company
                company_el = card.find(attrs={"data-testid": "job-item-company-name"})
                if not company_el:
                    company_el = card.find(class_=lambda x: x and "company" in str(x).lower())
                company = company_el.get_text(strip=True) if company_el else "N/A"

                # Location
                location_el = card.find(attrs={"data-testid": "job-item-location"})
                location = location_el.get_text(strip=True) if location_el else "Remote/DE"

                # Link
                link_el = card.find("a", href=True)
                link = "https://www.stepstone.de" + link_el["href"] if link_el and link_el["href"].startswith("/") else (link_el["href"] if link_el else url)

                job_id = f"stepstone_{title}_{company}".replace(" ", "_")[:80]
                jobs.append({
                    "id": job_id,
                    "source": "StepStone",
                    "title": title,
                    "company": company,
                    "location": location,
                    "link": link,
                    "query": query,
                })
            except Exception:
                continue
    except Exception as e:
        print(f"  StepStone error for '{query}': {e}")
    return jobs

# ─────────────────────────────────────────────────────────────────────────────
# SCRAPER: Indeed.de
# ─────────────────────────────────────────────────────────────────────────────
def scrape_indeed(query):
    jobs = []
    try:
        q = query.replace(" ", "+")
        url = f"https://de.indeed.com/jobs?q={q}&l=Deutschland&sc=0kf%3Aattr%28DSQF7%29%3B&remotejob=032b3046-06a3-4876-8dfd-474eb5e7ed11"
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            # Try without remote filter
            url = f"https://de.indeed.com/jobs?q={q}&l=remote"
            resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            return jobs

        soup = BeautifulSoup(resp.text, "html.parser")
        cards = soup.find_all("div", class_=lambda x: x and "job_seen_beacon" in str(x))
        if not cards:
            cards = soup.find_all("div", attrs={"data-testid": "slider_item"})
        if not cards:
            cards = soup.find_all("td", class_="resultContent")

        for card in cards[:8]:
            try:
                title_el = card.find(["h2", "span"], attrs={"title": True})
                if not title_el:
                    title_el = card.find(["h2", "a"])
                title = title_el.get_text(strip=True) if title_el else "N/A"

                company_el = card.find(attrs={"data-testid": "company-name"})
                if not company_el:
                    company_el = card.find(class_=lambda x: x and "companyName" in str(x))
                company = company_el.get_text(strip=True) if company_el else "N/A"

                location_el = card.find(attrs={"data-testid": "text-location"})
                if not location_el:
                    location_el = card.find(class_=lambda x: x and "companyLocation" in str(x))
                location = location_el.get_text(strip=True) if location_el else "Remote/DE"

                link_el = card.find("a", href=True)
                if link_el:
                    href = link_el["href"]
                    link = "https://de.indeed.com" + href if href.startswith("/") else href
                else:
                    link = url

                job_id = f"indeed_{title}_{company}".replace(" ", "_")[:80]
                jobs.append({
                    "id": job_id,
                    "source": "Indeed.de",
                    "title": title,
                    "company": company,
                    "location": location,
                    "link": link,
                    "query": query,
                })
            except Exception:
                continue
    except Exception as e:
        print(f"  Indeed error for '{query}': {e}")
    return jobs

# ─────────────────────────────────────────────────────────────────────────────
# SCRAPER: LinkedIn Jobs (public search)
# ─────────────────────────────────────────────────────────────────────────────
def scrape_linkedin(query):
    jobs = []
    try:
        q = query.replace(" ", "%20")
        url = f"https://www.linkedin.com/jobs/search/?keywords={q}&location=Germany&f_WT=2&sortBy=DD"
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            return jobs

        soup = BeautifulSoup(resp.text, "html.parser")
        cards = soup.find_all("div", class_=lambda x: x and "base-card" in str(x))
        if not cards:
            cards = soup.find_all("li", class_=lambda x: x and "result-card" in str(x))

        for card in cards[:8]:
            try:
                title_el = card.find(["h3", "span"], class_=lambda x: x and "title" in str(x).lower())
                title = title_el.get_text(strip=True) if title_el else "N/A"

                company_el = card.find(["h4", "a"], class_=lambda x: x and "company" in str(x).lower())
                company = company_el.get_text(strip=True) if company_el else "N/A"

                location_el = card.find("span", class_=lambda x: x and "location" in str(x).lower())
                location = location_el.get_text(strip=True) if location_el else "Remote/DE"

                link_el = card.find("a", href=True)
                link = link_el["href"].split("?")[0] if link_el else url

                job_id = f"linkedin_{title}_{company}".replace(" ", "_")[:80]
                jobs.append({
                    "id": job_id,
                    "source": "LinkedIn",
                    "title": title,
                    "company": company,
                    "location": location,
                    "link": link,
                    "query": query,
                })
            except Exception:
                continue
    except Exception as e:
        print(f"  LinkedIn error for '{query}': {e}")
    return jobs

# ─────────────────────────────────────────────────────────────────────────────
# EXTRACT COMPANY DOMAIN (for LinkedIn outreach)
# ─────────────────────────────────────────────────────────────────────────────
def get_linkedin_search_url(company):
    q = company.replace(" ", "%20")
    return f"https://www.linkedin.com/search/results/companies/?keywords={q}"

# ─────────────────────────────────────────────────────────────────────────────
# BUILD EMAIL HTML
# ─────────────────────────────────────────────────────────────────────────────
def build_email_html(new_jobs, date_str):
    if not new_jobs:
        return None

    # Group by source
    by_source = {}
    for job in new_jobs:
        src = job["source"]
        if src not in by_source:
            by_source[src] = []
        by_source[src].append(job)

    # Outreach template snippet
    outreach = (
        "Hallo [Name], ich habe gesehen dass ihr gerade einen [Position] sucht. "
        "Bevor ihr einen deutschen Vollzeitmitarbeiter einstellt — wäre es einen kurzen Austausch wert? "
        "Wir stellen deutschsprachige Remote-Mitarbeiter aus Kosovo zur Verfügung, 40h/Woche, "
        "zu ca. 50% der deutschen Personalkosten. Gleiche Zeitzone, gleiches Qualitätsniveau. "
        "Hättest du 10 Minuten?"
    )

    html = f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
  body {{ font-family: Arial, sans-serif; background: #f4f4f4; margin: 0; padding: 20px; }}
  .container {{ max-width: 700px; margin: 0 auto; background: white; border-radius: 8px; overflow: hidden; }}
  .header {{ background: #0D1B3E; color: white; padding: 24px 28px; }}
  .header h1 {{ margin: 0; font-size: 20px; }}
  .header p {{ margin: 6px 0 0; color: #C9A84C; font-size: 13px; }}
  .summary {{ background: #C9A84C; color: #0D1B3E; padding: 12px 28px; font-weight: bold; font-size: 14px; }}
  .section {{ padding: 20px 28px 0; }}
  .section h2 {{ color: #0D1B3E; font-size: 15px; border-bottom: 2px solid #C9A84C; padding-bottom: 6px; margin-bottom: 14px; }}
  .job-card {{ background: #f9f9f9; border-left: 4px solid #C9A84C; padding: 12px 14px; margin-bottom: 10px; border-radius: 0 6px 6px 0; }}
  .job-title {{ font-weight: bold; color: #0D1B3E; font-size: 14px; margin: 0 0 4px; }}
  .job-company {{ color: #333; font-size: 13px; margin: 0 0 2px; }}
  .job-meta {{ color: #777; font-size: 12px; margin: 0 0 8px; }}
  .job-links {{ font-size: 12px; }}
  .job-links a {{ color: #0D1B3E; margin-right: 14px; text-decoration: none; font-weight: bold; }}
  .job-links a:hover {{ text-decoration: underline; }}
  .outreach {{ background: #EEF1F7; border: 1px solid #CBD5E8; padding: 14px 16px; margin: 20px 28px; border-radius: 6px; font-size: 13px; color: #333; }}
  .outreach strong {{ color: #0D1B3E; display: block; margin-bottom: 8px; }}
  .footer {{ background: #0D1B3E; color: #aaa; padding: 16px 28px; font-size: 12px; margin-top: 20px; }}
</style>
</head>
<body>
<div class="container">
  <div class="header">
    <h1>A&amp;E Consulting — Daily Job Alert</h1>
    <p>{date_str} · {len(new_jobs)} new opportunities found</p>
  </div>
  <div class="summary">
    🎯 {len(new_jobs)} new job postings match your target positions today
  </div>
"""

    for source, jobs in by_source.items():
        html += f"""
  <div class="section">
    <h2>📌 {source} — {len(jobs)} posting{'s' if len(jobs) > 1 else ''}</h2>
"""
        for job in jobs:
            linkedin_url = get_linkedin_search_url(job['company'])
            html += f"""
    <div class="job-card">
      <p class="job-title">{job['title']}</p>
      <p class="job-company">🏢 {job['company']}</p>
      <p class="job-meta">📍 {job['location']} &nbsp;|&nbsp; 🔍 Query: <em>{job['query']}</em></p>
      <div class="job-links">
        <a href="{job['link']}" target="_blank">→ View Job Posting</a>
        <a href="{linkedin_url}" target="_blank">→ Find on LinkedIn</a>
      </div>
    </div>
"""
        html += "  </div>"

    html += f"""
  <div class="outreach">
    <strong>💬 Ready-to-use LinkedIn outreach template:</strong>
    {outreach}
  </div>
  <div class="footer">
    A&amp;E Consulting LLC · enes@aeconsultingllc.de · aeconsultingllc.de<br>
    This digest is generated automatically every morning. Reply rate tip: message within 24h of posting going live.
  </div>
</div>
</body>
</html>
"""
    return html


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL LOGGER — saves all jobs to ae_job_leads.xlsx
# ─────────────────────────────────────────────────────────────────────────────
def save_to_excel(new_jobs):
    filepath = CONFIG["excel_file"]
    date_str = datetime.now().strftime("%d.%m.%Y")

    # Colors
    NAVY   = "0D1B3E"
    GOLD   = "C9A84C"
    WHITE  = "FFFFFF"
    LIGHT  = "F5F5F5"
    GREEN  = "E8F5EE"

    headers = [
        "Date Found", "Source", "Job Title", "Company",
        "Location", "Search Query", "Job Link", "LinkedIn Search", "Contacted"
    ]
    col_widths = [14, 12, 35, 30, 20, 30, 50, 50, 12]

    # Load or create workbook
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
        next_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Leads"

        # Header row
        header_fill = PatternFill("solid", fgColor=NAVY)
        header_font = Font(bold=True, color=WHITE, size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            ws.column_dimensions[get_column_letter(col)].width = w

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        next_row = 2

    # Write new jobs
    for i, job in enumerate(new_jobs):
        row = next_row + i
        bg = LIGHT if row % 2 == 0 else WHITE

        values = [
            date_str,
            job.get("source", ""),
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            job.get("query", ""),
            job.get("link", ""),
            f"https://www.linkedin.com/search/results/companies/?keywords={job.get('company','').replace(' ', '%20')}",
            "No"
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.font = Font(size=10)

            # Contacted column — green background
            if col == 9:
                cell.fill = PatternFill("solid", fgColor="FFF9E6")
                cell.font = Font(size=10, bold=True, color="8B6914")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Hyperlinks for link columns
            if col in (7, 8) and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(size=10, color="0563C1", underline="single")

        ws.row_dimensions[row].height = 18

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    wb.save(filepath)
    print(f"  -> Excel saved: {filepath} ({len(new_jobs)} new rows)")
    return filepath

# ─────────────────────────────────────────────────────────────────────────────
# SEND EMAIL
# ─────────────────────────────────────────────────────────────────────────────
def send_email(html_body, job_count, date_str):
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"🎯 A&E Job Alert — {job_count} new opportunities — {date_str}"
        msg["From"]    = CONFIG["sender_email"]
        msg["To"]      = CONFIG["recipient_email"]
        msg.attach(MIMEText(html_body, "html"))

        with smtplib.SMTP(CONFIG["smtp_host"], CONFIG["smtp_port"]) as server:
            server.ehlo()
            server.starttls()
            server.login(CONFIG["sender_email"], CONFIG["sender_password"])
            server.sendmail(CONFIG["sender_email"], CONFIG["recipient_email"], msg.as_string())
        print(f"  ✅ Email sent to {CONFIG['recipient_email']}")
        return True
    except Exception as e:
        print(f"  ❌ Email error: {e}")
        print("  → Check sender_email and sender_password in CONFIG")
        return False

# ─────────────────────────────────────────────────────────────────────────────
# MAIN RUN
# ─────────────────────────────────────────────────────────────────────────────
def run():
    date_str = datetime.now().strftime("%d.%m.%Y %H:%M")
    print(f"\n{'='*60}")
    print(f"A&E Consulting — Job Monitor")
    print(f"Run time: {date_str}")
    print(f"{'='*60}")

    seen_jobs = load_seen_jobs()
    all_new_jobs = []
    scrapers = [
        ("StepStone", scrape_stepstone),
        ("Indeed.de", scrape_indeed),
        ("LinkedIn",  scrape_linkedin),
    ]

    for query in SEARCH_QUERIES:
        print(f"\n🔍 Searching: '{query}'")
        for source_name, scraper_fn in scrapers:
            print(f"  → {source_name}...", end=" ")
            jobs = scraper_fn(query)
            new = [j for j in jobs if j["id"] not in seen_jobs]
            print(f"{len(new)} new")
            for j in new:
                seen_jobs.add(j["id"])
                all_new_jobs.append(j)
            time.sleep(random.uniform(CONFIG["min_delay"], CONFIG["max_delay"]))

    print(f"\n{'='*60}")
    print(f"Total new jobs found: {len(all_new_jobs)}")

    if all_new_jobs:
        # Save to Excel first
        save_to_excel(all_new_jobs)
        # Send email digest
        html = build_email_html(all_new_jobs, date_str)
        if html:
            send_email(html, len(all_new_jobs), date_str)
            save_seen_jobs(seen_jobs)
            # Save HTML locally for review
            with open("last_digest.html", "w", encoding="utf-8") as f:
                f.write(html)
            print(f"  → Digest also saved to: last_digest.html")
    else:
        print("  → No new jobs found. Nothing sent.")
        save_seen_jobs(seen_jobs)

    print(f"{'='*60}\n")

# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # Check config
    run()

# ─────────────────────────────────────────────────────────────────────────────
# WINDOWS TASK SCHEDULER SETUP (run every day at 7:00 AM)
# ─────────────────────────────────────────────────────────────────────────────
"""
TO SCHEDULE DAILY AT 7:00 AM ON WINDOWS:

1. Open Task Scheduler (search in Start Menu)
2. Click "Create Basic Task"
3. Name: "AE Job Monitor"
4. Trigger: Daily at 07:00
5. Action: Start a program
   Program: C:/Users/hazir/AppData/Local/Programs/Python/Python311/python.exe
   Arguments: job_monitor_ae.py
   Start in: C:/path/to/folder/where/this/file/is
6. Finish

OR run this command in PowerShell once (replace paths):
schtasks /create /tn "AE Job Monitor" /tr "python C:/path/to/job_monitor_ae.py" /sc daily /st 07:00
"""
